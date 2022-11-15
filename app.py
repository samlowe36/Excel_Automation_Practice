import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):         #in this case transactions.xlsx
    wb = xl.load_workbook(filename)     #in this case transactions.xlsx
    sheet = wb["Sheet1"]
    #cell = sheet["a1"]     #sets the cell to a1 by default
    #cell = sheet.cell(1, 1)   #alternate way to do the same thing as above
    #print(cell.value)  #prints value in cell (in this case it is currently value in a1)
    #print(sheet.max_row) (prints the max number of rows (in this case it is 4)

    for row in range(2, sheet.max_row + 1):     #we start at 2 instead of 1 to skip the header
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")
    wb.save(filename)   #in this case transactions.xlsx